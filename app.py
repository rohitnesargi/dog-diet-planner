from flask import Flask, render_template, request, redirect, url_for, session, send_file
from werkzeug.utils import secure_filename
import os
import numpy as np
from PIL import Image
import tensorflow as tf
from tensorflow.keras.applications.mobilenet_v2 import preprocess_input, decode_predictions
import openpyxl
from diet_data import generate_diet_plan, get_size_category, get_exact_size
import datetime

app = Flask(__name__)
app.secret_key = "secretkey"

# ======================================================
# CONFIG
# ======================================================

UPLOAD_FOLDER = 'static/uploads'
EXCEL_FILE = 'users.xlsx'

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

ADMIN_EMAIL = "admin@gmail.com"
ADMIN_PASSWORD = "admin123"

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# ======================================================
# CREATE EXCEL FILE
# ======================================================

if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Users"
    sheet.append(["Username", "Email", "Password"])
    wb.save(EXCEL_FILE)

# ======================================================
# CREATE PET HEALTH EXCEL FILE
# ======================================================
HEALTH_FILE = 'pet_health.xlsx'

if not os.path.exists(HEALTH_FILE):
    wb = openpyxl.Workbook()
    
    # Vaccinations Sheet
    ws1 = wb.active
    ws1.title = "Vaccinations"
    ws1.append(["UserEmail", "PetName", "Vaccine", "Date", "Status"])
    
    # Grooming Sheet
    ws2 = wb.create_sheet("Grooming")
    ws2.append(["UserEmail", "PetName", "Service", "Date", "Time"])
    
    wb.save(HEALTH_FILE)

# Global model variable
model = None

def get_model():
    global model
    if model is None:
        print("Loading AI Model... This may take a moment.")
        model = tf.keras.applications.MobileNetV2(weights='imagenet')
    return model

# ======================================================
# USER FUNCTIONS
# ======================================================

def register_user(username, email, password):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active
        
        email_to_check = email.strip().lower()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            db_email = str(row[1]).strip().lower() if row[1] is not None else ""
            if db_email == email_to_check:
                return False

        sheet.append([
            username.strip(),
            email_to_check,
            password
        ])

        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        print(f"Error in register_user: {e}")
        return False


def check_user(email, password):
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            
            db_user = str(row[0]).strip() if row[0] is not None else ""
            db_email = str(row[1]).strip().lower() if row[1] is not None else ""
            db_pass = str(row[2]).strip() if row[2] is not None else ""

            if db_email == email.strip().lower() and db_pass == password.strip():
                return db_user if db_user else "User"

        return None
    except Exception as e:
        print(f"Error in check_user: {e}")
        return None

# ======================================================
# PREDICT BREED
# ======================================================

def predict_breed(img_path):

    img = Image.open(img_path).convert('RGB').resize((224, 224))

    img_array = np.array(img)

    img_array = preprocess_input(img_array)

    img_array = np.expand_dims(img_array, axis=0)

    model_instance = get_model()
    predictions = model_instance.predict(img_array)

    decoded = decode_predictions(predictions, top=1)[0][0]

    breed = str(decoded[1])

    confidence = int(round(decoded[2] * 100))

    return breed, confidence

# ======================================================
# SERVE PHOTOS
# ======================================================

@app.route('/photos/<filename>')
def serve_photos(filename):
    return send_file(os.path.join('photos', filename))

# ======================================================
# HOME
# ======================================================

@app.route('/')
def home():
    return redirect(url_for('login'))

# ======================================================
# REGISTER
# ======================================================

@app.route('/register', methods=['GET', 'POST'])
def register():

    if request.method == 'POST':

        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')

        if not username:
            return render_template('register.html', message="Username is required!")
        if not email:
            return render_template('register.html', message="Email is required!")
        if not password:
            return render_template('register.html', message="Password is required!")

        success = register_user(username, email, password)

        if not success:
            return render_template('register.html',
                                   message="Email already exists!")

        return render_template('login.html',
                               message="Registration Successful! Please Login")

    return render_template('register.html')

# ======================================================
# LOGIN
# ======================================================

@app.route('/login', methods=['GET', 'POST'])
def login():

    if request.method == 'POST':

        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')

        if not email:
            return render_template('login.html',
                                   message="Email is required!")
        if not password:
            return render_template('login.html',
                                   message="Password is required!")

        # Admin Login
        if email == ADMIN_EMAIL and password == ADMIN_PASSWORD:
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))

        # User Login
        user = check_user(email, password)

        if user:
            session['user'] = user
            return redirect(url_for('index'))

        return render_template('login.html',
                               message="Invalid Credentials")

    return render_template('login.html')

# ======================================================
# LOGOUT
# ======================================================

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ======================================================
# MAIN PAGE
# ======================================================

@app.route('/index', methods=['GET', 'POST'])
def index():

    if 'user' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':

        if 'image' not in request.files:
            return render_template('index.html',
                                   user=session['user'],
                                   message="Upload image")

        file = request.files['image']

        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        file.save(filepath)

        breed, confidence = predict_breed(filepath)

        size = str(get_size_category(breed))
        exact_size = get_exact_size(breed, size)

        diet = generate_diet_plan(breed, size)

        youtube_link = f"https://www.youtube.com/results?search_query={breed}+dog+training"

        amazon_link = "https://www.amazon.in/s?k=" + diet['food'].replace(" ", "+")

        session['breed'] = str(breed)
        session['confidence'] = float(confidence)
        session['size'] = str(size)
        session['exact_size'] = exact_size
        session['diet'] = dict(diet)

        # Pass the relative path for use with url_for in the template
        # 'static/uploads/filename' -> 'uploads/filename'
        relative_image_path = os.path.join('uploads', filename).replace('\\', '/')

        return render_template(
            'result.html',
            image=relative_image_path,
            breed=breed,
            confidence=confidence,
            diet=diet,
            size=size,
            exact_size=exact_size,
            youtube_link=youtube_link,
            amazon_link=amazon_link,
            user=session['user']
        )

    return render_template('index.html', user=session['user'])

# ======================================================
# DOWNLOAD REPORT
# ======================================================

@app.route('/download_report')
def download_report():

    if 'user' not in session:
        return redirect(url_for('login'))

    report = f"""
DOG REPORT
====================

Breed: {session['breed']}
Confidence: {session['confidence']} %

Size: {session['size']}

Food: {session['diet']['food']}
Meals: {session['diet']['meals']}
Extras: {session['diet']['extras']}

Generated:
{datetime.datetime.now()}
"""

    file_path = "dog_report.txt"

    with open(file_path, "w") as file:
        file.write(report)

    return send_file(file_path, as_attachment=True)

# ======================================================
# ADMIN LOGIN
# ======================================================

@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():

    if request.method == 'POST':

        email = request.form.get('email')
        password = request.form.get('password')

        if email == ADMIN_EMAIL and password == ADMIN_PASSWORD:
            session['admin'] = True
            return redirect(url_for('admin_dashboard'))

        return render_template("admin_login.html",
                               message="Invalid Admin Credentials")

    return render_template("admin_login.html")

# ======================================================
# ADMIN DASHBOARD
# ======================================================

@app.route('/admin')
def admin_dashboard():

    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    users = []

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        users.append({
            "id": i,
            "username": row[0],
            "email": row[1],
            "password": row[2]
        })

    return render_template("admin.html", users=users)

# ======================================================
# DELETE USER
# ======================================================

@app.route('/delete_user/<int:user_id>')
def delete_user(user_id):

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    sheet.delete_rows(user_id + 2)

    wb.save(EXCEL_FILE)

    return redirect(url_for('admin_dashboard'))

# ======================================================
# EDIT USER
# ======================================================

@app.route('/edit_user/<int:user_id>', methods=['GET', 'POST'])
def edit_user(user_id):

    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active

    row_no = user_id + 2

    if request.method == 'POST':

        sheet.cell(row=row_no, column=1).value = request.form['username']
        sheet.cell(row=row_no, column=2).value = request.form['email']
        sheet.cell(row=row_no, column=3).value = request.form['password']

        wb.save(EXCEL_FILE)

        return redirect(url_for('admin_dashboard'))

    user = {
        "username": sheet.cell(row=row_no, column=1).value,
        "email": sheet.cell(row=row_no, column=2).value,
        "password": sheet.cell(row=row_no, column=3).value
    }

    return render_template("edit_user.html", user=user)

# ======================================================
# ADMIN LOGOUT
# ======================================================

@app.route('/admin_logout')
def admin_logout():
    session.pop('admin', None)
    return redirect(url_for('login'))

# ======================================================
# EXTRA PAGES
# ======================================================

@app.route('/about')
def about():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('about.html', user=session['user'])

@app.route('/services')
def services():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('services.html', user=session['user'])

@app.route('/vets')
def vets():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('vets.html', user=session['user'])

@app.route('/profile')
def profile():
    if 'user' not in session:
        return redirect(url_for('login'))
    return render_template('profile.html', user=session['user'])

@app.route('/contact', methods=['GET', 'POST'])
def contact():
    if 'user' not in session:
        return redirect(url_for('login'))
    if request.method == 'POST':
        return render_template('contact.html', user=session['user'], message="Message sent successfully!")
    return render_template('contact.html', user=session['user'])

# ======================================================
# NEW MODULES
# ======================================================

# DISEASE PREDICTION
@app.route('/disease_prediction', methods=['GET', 'POST'])
def disease_prediction():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    result = None
    if request.method == 'POST':
        symptoms = request.form.getlist('symptoms')
        
        # Simple Logic
        if set(['coughing', 'lethargy']).issubset(set(symptoms)):
            result = {"disease": "Kennel Cough", "advice": "Consult a vet for antibiotics and keep your dog isolated."}
        elif set(['vomiting', 'diarrhea', 'lethargy']).issubset(set(symptoms)):
            result = {"disease": "Parvovirus", "advice": "URGENT: This is life-threatening. Visit an emergency vet immediately."}
        elif set(['fever', 'sneezing', 'coughing']).issubset(set(symptoms)):
            result = {"disease": "Distemper", "advice": "Consult a vet. This is a serious viral disease."}
        elif 'itching' in symptoms:
            result = {"disease": "Allergies", "advice": "Check for skin parasites or change in diet. Consult a vet for antihistamines."}
        elif symptoms:
            result = {"disease": "General Infection/Malaise", "advice": "Your dog seems unwell. Rest and monitor, but consult a vet if symptoms persist."}
        else:
            result = {"disease": "Unknown", "advice": "Please select some symptoms for a better analysis."}

    return render_template('disease_prediction.html', user=session['user'], result=result)

# VACCINATION REMINDER
@app.route('/vaccination', methods=['GET', 'POST'])
def vaccination():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    user_email = ""
    # Get user email from session/excel (need to store email in session for easier access)
    # Re-checking user email from users.xlsx based on session['user']
    wb_u = openpyxl.load_workbook(EXCEL_FILE)
    sh_u = wb_u.active
    for row in sh_u.iter_rows(min_row=2, values_only=True):
        if row[0] == session['user']:
            user_email = row[1]
            break

    if request.method == 'POST':
        pet_name = request.form.get('pet_name')
        vaccine = request.form.get('vaccine')
        date = request.form.get('date')
        
        wb = openpyxl.load_workbook(HEALTH_FILE)
        sheet = wb["Vaccinations"]
        sheet.append([user_email, pet_name, vaccine, date, "Scheduled"])
        wb.save(HEALTH_FILE)
        return redirect(url_for('vaccination'))

    # Load vaccinations
    vaccinations = []
    wb = openpyxl.load_workbook(HEALTH_FILE)
    sheet = wb["Vaccinations"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user_email:
            vaccinations.append({"pet": row[1], "vaccine": row[2], "date": row[3], "status": row[4]})

    return render_template('vaccination.html', user=session['user'], vaccinations=vaccinations)

# GROOMING SCHEDULER
@app.route('/grooming', methods=['GET', 'POST'])
def grooming():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_email = ""
    wb_u = openpyxl.load_workbook(EXCEL_FILE)
    sh_u = wb_u.active
    for row in sh_u.iter_rows(min_row=2, values_only=True):
        if row[0] == session['user']:
            user_email = row[1]
            break

    if request.method == 'POST':
        pet_name = request.form.get('pet_name')
        service = request.form.get('service')
        date = request.form.get('date')
        time = request.form.get('time')
        
        wb = openpyxl.load_workbook(HEALTH_FILE)
        sheet = wb["Grooming"]
        sheet.append([user_email, pet_name, service, date, time])
        wb.save(HEALTH_FILE)
        return redirect(url_for('grooming'))

    grooming_sessions = []
    wb = openpyxl.load_workbook(HEALTH_FILE)
    sheet = wb["Grooming"]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user_email:
            grooming_sessions.append({"pet": row[1], "service": row[2], "date": row[3], "time": row[4]})

    return render_template('grooming.html', user=session['user'], sessions=grooming_sessions)

# HEALTH DASHBOARD
@app.route('/health_dashboard')
def health_dashboard():
    if 'user' not in session:
        return redirect(url_for('login'))

    user_email = ""
    wb_u = openpyxl.load_workbook(EXCEL_FILE)
    sh_u = wb_u.active
    for row in sh_u.iter_rows(min_row=2, values_only=True):
        if row[0] == session['user']:
            user_email = row[1]
            break

    wb = openpyxl.load_workbook(HEALTH_FILE)
    
    # Latest Vaccine
    vaccines = []
    for row in wb["Vaccinations"].iter_rows(min_row=2, values_only=True):
        if row[0] == user_email:
            vaccines.append({"pet": row[1], "vaccine": row[2], "date": row[3]})
    
    # Latest Grooming
    grooming = []
    for row in wb["Grooming"].iter_rows(min_row=2, values_only=True):
        if row[0] == user_email:
            grooming.append({"pet": row[1], "service": row[2], "date": row[3]})

    return render_template('health_dashboard.html', user=session['user'], vaccines=vaccines, grooming=grooming)

# AI CHATBOT
@app.route('/chatbot', methods=['GET', 'POST'])
def chatbot():
    if 'user' not in session:
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        user_msg = request.json.get('message', '').lower()
        
        response = "I'm sorry, I don't understand that. You can ask me about diet, symptoms, or grooming!"
        
        if 'diet' in user_msg or 'food' in user_msg:
            response = "For a healthy diet, ensure your dog gets balanced proteins and vitamins. You can use our Diet Planner for specific breed recommendations!"
        elif 'fever' in user_msg or 'sick' in user_msg or 'symptoms' in user_msg:
            response = "If your dog is showing symptoms like fever or lethargy, please use our Disease Prediction tool or consult a vet immediately."
        elif 'groom' in user_msg or 'bath' in user_msg:
            response = "Regular grooming is essential for a healthy coat. You can schedule a session in our Grooming Scheduler!"
        elif 'hi' in user_msg or 'hello' in user_msg:
            response = "Hello! I'm your AI Dog Care Assistant. How can I help you today?"
            
        return {"response": response}

    return render_template('chatbot.html', user=session['user'])

# ======================================================

if __name__ == '__main__':
    app.run(debug=True)
